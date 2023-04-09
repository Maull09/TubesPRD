import openpyxl
import serial
import datetime
import tkinter as tk


wb = openpyxl.load_workbook('data.xlsx')
mahasiswa_ws = wb['Mahasiswa']
kelas_ws = wb['Kelas']
kehadiran_ws = None


ser = serial.Serial('COM10', 9600) # sesuaikan dengan port serial pada komputer Anda
last_scan = None


def get_current_kelas():
    current_time = datetime.datetime.now().strftime('%H:%M:%S')
    current_date = datetime.datetime.now().strftime('%d-%m-%Y')
    
    kelas_data = {}

    # Check if there is a class scheduled at the current time and date
    for row in kelas_ws.iter_rows(min_row=2, values_only=True):
        start_time_str = row[5].strftime('%H:%M:%S')
        end_time_str = row[6].strftime('%H:%M:%S')
        date_class_str = row[7].strftime('%d-%m-%Y')

        if date_class_str == current_date and start_time_str <= current_time <= end_time_str:
            kode_kelas = row[0]
            kode_mata_kuliah = row[1]
            lokasi = row[2]
            mata_kuliah = row[3]
            dosen = row[4]
            kelas_data = {
                'kode_kelas': kode_kelas,
                'kode_mata_kuliah': kode_mata_kuliah,
                'lokasi': lokasi,
                'mata_kuliah': mata_kuliah,
                'dosen': dosen
            }
            break

    return kelas_data


current_time = datetime.datetime.now().strftime('%H:%M:%S')
current_date = datetime.datetime.now().strftime('%d-%m-%Y')


if current_date in wb.sheetnames:
    kehadiran_ws = wb[current_date]
else:
    kehadiran_ws = wb.create_sheet(current_date)
    kehadiran_ws.append(['Waktu', 'Nama', 'NIM', 'Tag RFID', 'Kehadiran', 'Mata Kuliah', 'Kode Mata Kuliah', 'Kode Kelas', 'Lokasi', 'Dosen'])


mahasiswa_data = []
for row in mahasiswa_ws.iter_rows(min_row=2, values_only=True):
    mahasiswa_data.append(row)


def scan_card():
    global last_scan
    data = ser.readline().decode().strip()

    if data and data != last_scan:
        last_scan = data
        name = None
        nim = None
        kehadiran = None
        mata_kuliah = None
        kode_mata_kuliah = None
        kode_kelas = None
        lokasi = None
        dosen = None

        # Find student data from Mahasiswa sheet
        for row in mahasiswa_data:
            if row[2] == data:
                name = row[0]
                nim = row[1]
                break
        
        # Find class data from Kelas sheet
        if name is not None:
            current_kelas = get_current_kelas()

            if current_kelas is not None :
                kode_kelas = current_kelas.get('kode_kelas')
                lokasi = current_kelas.get('lokasi')
                mata_kuliah = current_kelas.get('mata_kuliah')
                kode_mata_kuliah = current_kelas.get('kode_mata_kuliah')
                dosen = current_kelas.get('dosen')
                kehadiran = 'Hadir'
        
        # Write data to Kehadiran sheet
        if name is not None:
            kehadiran_ws.append([f'{current_date} {current_time}', name, nim, data, kehadiran, mata_kuliah, kode_mata_kuliah, kode_kelas, lokasi, dosen])
            wb.save('data.xlsx')
            gedung.config(text=f"{lokasi}")
            nama.config(text=f"Nama                      : {name}")
            nimm.config(text=f"NIM                         : {nim}")
            status.config(text=f"Status Kehadiran    : {kehadiran}" )
            mata_kuliahh.config(text=f"{mata_kuliah}")
            kode_kelass.config(text=f"{kode_kelas}")

        else:
            print(f'Data not found! {data}')

    root.after(1000,scan_card)

# Create GUI
root = tk.Tk()
root.title("Mesin Absensi dan Rekap Kehadiran (MARK)")
root.geometry("1920x1080")

# Set background image
bg_image = tk.PhotoImage(file="itboy.png")
bg_label = tk.Label(root, image=bg_image)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# Create labels to display attendance data
nama = tk.Label(root, text="Nama : Mohamad Maulana Firdaus Ramadhan", font=("Helvetica", 18), bg="#ffffff")
nama.place(x=400, y=600)

nimm = tk.Label(root, text="NIM    : 19622267", font=("Helvetica", 18), bg="#ffffff")
nimm.place(x=400, y=650)

status = tk.Label(root, text="Status: Hadir", font=("Helvetica", 18), bg="#ffffff")
status.place(x=400, y=700)

gedung = tk.Label(root, text="K2. 9653", font=("Helvetica", 24), bg="#ffffff")
gedung.place(x=728, y=100)

mata_kuliahh = tk.Label(root, text="Kalkulus", font=("Helvetica", 20), bg="#ffffff")
mata_kuliahh.place(x=742, y=150)

kode_kelass = tk.Label(root, text="K-51", font=("Helvetica", 20), bg="#ffffff")
kode_kelass.place(x=764, y=200)

root.after(2000,scan_card)

root.mainloop()